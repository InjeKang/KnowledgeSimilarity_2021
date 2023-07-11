from cmath import isnan
import multiprocessing as mp
from tokenize import PseudoExtras
import pandas as pd
import os
import openpyxl
import glob
from tqdm import trange
import math
from multiprocessing import Pool, cpu_count
import numpy as np
from functools import partial
from os.path import join
import re


def match_ctrl_var(ally_data, firm_data): #raw_, data
    city_ = []
    state_ = []
    employ_ = []
    RnD_ = []
    revenue_ = []
    age_ = []
    semi_ind_ = []
    for i in range(len(ally_data)):
        # two dataframe to find one
        # (1) that matches only the firm (needed to calculate 'age') and
        # (2) that matches both firm and year
        # focal firms' characteristics
        firm_raw = firm_data[firm_data["Company Name"].str.contains(ally_data["focal"].iloc[i].lower())].reset_index()     
        raw = firm_data[(firm_data["Company Name"].str.contains(ally_data["focal"].iloc[i].lower())) & 
                        (firm_data["Data Year - Fiscal"] == ally_data["year"].iloc[i])].reset_index()        
        if raw.empty: # to consider alliance data that does not match with firm data with regards "year"
            city_.append("") # when use "extend", the empty string is not included
            state_.append("")
            employ_.append("")
            RnD_.append("")
            revenue_.append("")
            age_.append("")
        else:
            city_.append(raw["City"].item())
            state_.append(raw["State/Province"].item())
            employ_.append(raw["Employees"].item())
            RnD_.append(raw["Research & Development - Prior"].item())
            revenue_.append(raw["Revenue - Total"].item())
            if math.isnan(firm_raw["IPO_Year"][0].item()): # because there is empty variables of which type is float nan
                age_.append("")
            else:
                age_.append(ally_data["year"][i] - int(firm_raw["IPO_Year"][0].item()))
    # if a partner firm is also a semiconductor industry
    semi_firm = list(set(ally_data["focal"].str.lower().tolist()))
    semi_ind_ = ally_data["partner"].str.lower().str.contains("|".join(semi_firm), case=False)
    return city_, state_, employ_, RnD_, revenue_, age_, semi_ind_

def add_ctrl_var(ally_data, firm_data):
    data = ally_data.copy()
    city, state, employee, RnD, revenue, age, semi_ind = match_ctrl_var(ally_data, firm_data)
    data["city"] = city
    data["state"] = state
    data["employ"] = employee
    data["RnD"] = RnD
    data["revenue"] = revenue
    data["age"] = age
    data["semi_ind"] = semi_ind
    return data



def read_xlsx(data):
    os.chdir("D:/Analysis/2021_Similarity/data/patent/KISTI")
    patent = openpyxl.load_workbook(data) # can no longer open xlsx file from pandas' read_excel >> used openpyxl
    if patent.sheetnames == ["Sheet0"]:
        sheet = patent["Sheet0"]
    elif patent.sheetnames == ["Sheet1"]:
        sheet = patent["Sheet1"]
    sheet.delete_rows(0, 2)
    firm_patent = pd.DataFrame(sheet.values)
    firm_patent = firm_patent[[9, 10, 18, 32]] #select columns - application number, application year, ipc
    os.chdir("D:/Analysis/2021_Similarity/data")
    return firm_patent

def read_xlsx_directory(input_path):    
    """ read all patents in a file"""
    # set directory
    os.chdir(input_path)
    # read all excel files in the directory
    excel_files = glob.glob((input_path + "\patent\KISTI\*.xlsx"))
    patent_stock = pd.concat((read_xlsx_glob(files) for files in excel_files))  
    return patent_stock.reset_index()  

def read_xlsx_glob(data):    
    patent = openpyxl.load_workbook(data) # can no longer open xlsx file from pandas' read_excel >> used openpyxl
    if patent.sheetnames == ["Sheet0"]:
        sheet = patent["Sheet0"]
    elif patent.sheetnames == ["Sheet1"]:
        sheet = patent["Sheet1"]
    sheet.delete_rows(0, 2)
    firm_patent = pd.DataFrame(sheet.values)
    firm_patent = firm_patent[[9, 10, 18, 32]] #select columns - application number, application year, assignee, ipc
    firmName = fileName(data)
    firm_patent["firm"] = firmName
    return firm_patent  

def fileName(x):
    stopwords = "D:\\Analysis\\2021_Similarity\\data\\patent\\KISTI\\", ".xlsx"
    stopwords_ = "|".join(map(re.escape, stopwords))
    x2 = re.sub(stopwords_, "", x)
    return x2

def split_column_into_lists(data, column, splitter):
    """Turn every rows of the column into Series of lists"""    
    column2 = column.copy()
    for i in range(len(column2)):
        try:
            column2.iloc[i] = str(column2.iloc[i]).split(splitter)
            for j in range(len(column2.iloc[i])):
                column2.iloc[i][j] = column2.iloc[i][j].strip()
        except:
            if isinstance(column2.iloc[i], str): # some columns consist of rows with list while others with string
                column2[i] = str(column2.iloc[i]).split(splitter)
                for j in range(len(column2.iloc[i])):
                    column2.iloc[i][j] = column2.iloc[i][j].strip()
            # else:
            #     # column2.iloc[i] = column2.iloc[i]
            #     pass
    column2list = column2
    return column2list

def column_to_list(data, column):
    """Make rows of a column into one list"""    
    if column.dropna().empty:
        firm_patent_list = []
    else:
        patent_list = split_column_into_lists(data,column, ";") # type(patent_list) = Series with lists
        # Series with lists to a list
        firm_patent_list = patent_list.apply(pd.Series).stack().reset_index(drop = True).tolist()
    return firm_patent_list


def create_placeholder_list(list_of_list):
    """Create a placeholder list (a) to prevent error of ind out of range and
    (b) to have a word (instead of an alphabet) as a string"""
    if isinstance(list_of_list, list):
        placeholder_list = []
        for i, _ in enumerate(list_of_list):
            placeholder_list_row = []
            for j, _ in enumerate(list_of_list[i]):
                placeholder_list_row.append("")
            placeholder_list.append(placeholder_list_row)
        return placeholder_list
    else:
        placeholder_list = []
        for i, _ in enumerate(list_of_list):
            placeholder_list_row = []
            for j, _ in enumerate(list_of_list.iloc[i]):
                placeholder_list_row.append("")
            placeholder_list.append(placeholder_list_row)
        return placeholder_list


def flatten_ipc(data):        
    flatten_list = []
    if isinstance(data, list):
        for i in range(len(data)):
            if isinstance(data[i], list): # lists in list
                data[i] = [strX for strX in data[i] if strX.strip()] # remove blank string
                data[i] = list((strX.replace(" ", "") for strX in data[i])) # remove whitespace
                data[i] = [split_(strX, ";") for strX in data[i]]
                data[i] = [strX for strX_list in data[i] for strX in strX_list] # to flatten a list of lists
                flatten_list.extend(data[i])                      
            elif isinstance(data[i], str):
                data[i] = data[i].replace(" ","")
                data[i] = split_(data[i], ";")
                flatten_list.extend(data[i])
            elif (data[i] == [] or data[i] == "") :
                flatten_list = data[i]
            elif ((data[i] is None) or math.isnan(data[i])):
                pass
            else:
                print(data[i])
    elif math.isnan(data):
        flatten_list = data
    else:
        print(data)
    return flatten_list

def ipc_diversity(firm, year, patent):
    firm = firm.lower()
    firm_patent = patent[(pd.to_numeric(patent[10]) >= year-5) & 
            (pd.to_numeric(patent[10]) <= year-1) &
            (patent["firm"].str.lower() == firm)].reset_index(drop=True, inplace=False)    
    firm_ipc = flatten_ipc(firm_patent[32].tolist())
    output = len(set(firm_ipc))
    return output


def patent_size(firm, year, patent):
    firm = firm.lower()
    firm_patent = patent[(pd.to_numeric(patent[10]) >= year-5) & 
            (pd.to_numeric(patent[10]) <= year-1) &
            (patent["firm"].str.lower() == firm)].reset_index(drop=True, inplace=False)    
    output = len(firm_patent)
    return output 


def presample(focal, partner, patent):
    """
    10 years of presample (i.e. ~ 1989 or 1990 ~ 1999)
    >> alliance data from 1990
    >> joint patents before alliance data (Runge et al., 2021)
    """
    patent2 = patent[(pd.to_numeric(patent[10]) <= 1999)].reset_index(drop=True, inplace=False)
                    # (pd.to_numeric(patent[10]) >= 1990)]
    # removing duplicated data
    # patent2 = patent2.drop_duplicates(subset = [9]).reset_index(drop=True, inplace=False)
    # retrieving patent assigned by both firms
    patent3 = patent2[(patent2["firm"].str.lower() == focal.lower()) |
                (patent2["firm"].str.lower() == partner.lower())]
    # return len(patent3)
    return patent3.duplicated(subset=[9]).sum()/10


def add_ctrl_var3(data, patent_stock):
    data2 = data.copy()
    data2["presample"] = data2.swifter.apply(lambda x: 
                                np.nan if math.isnan(x["tech_sim"])
                                else presample(x["focal"], x["partner"], patent_stock), axis=1)
    data2["no_ipc_focal"] = data2.swifter.apply(lambda x:
                                np.nan if math.isnan(x["tech_sim"])
                                else ipc_diversity(x["focal"], x["year"], patent_stock), axis=1)
    data2["no_ipc_partner"] = data2.swifter.apply(lambda x:
                                np.nan if math.isnan(x["tech_sim"])
                                else ipc_diversity(x["partner"], x["year"], patent_stock), axis=1)
    data2["patent_size_focal"] = data2.swifter.apply(lambda x:
                                np.nan if math.isnan(x["tech_sim"])
                                else patent_size(x["focal"], x["year"], patent_stock), axis=1)
    data2["patent_size_partner"] = data2.swifter.apply(lambda x:
                                np.nan if math.isnan(x["tech_sim"])
                                else patent_size(x["partner"], x["year"], patent_stock), axis=1)
    return data2


def multi_process(df, target_func, type_): # type_ = df or list
    n_cores = 12
    if type_ ==  "df":        
        df_split = np.array_split(df, n_cores)
        pool = Pool(n_cores)
        output = pd.concat(pool.map(target_func, df_split))
    else:
        list_ = []        
        list_split = np.array_split(df, n_cores)
        pool = Pool(n_cores)
        output = list_.append(pool.map(target_func, list_split))
    """
    When multiprocessing dataframe, check if .iloc is properly used to prevent KeyError: 0
        
    pool.apply: the function call is performed in a seperate process / blocks until the function is completed / lack of reducing time
    pool.apply_async: returns immediately instead of waiting for the result / the orders are not the same as the order of the calls
    pool.map: list of jobs in one time (concurrence) / block / ordered-results
    pool.map_async: 
    http://blog.shenwei.me/python-multiprocessing-pool-difference-between-map-apply-map_async-apply_async/
    """
    pool.close()
    pool.join()
    return output




def split_(data, splitter):
    """to split string with the consideration of NoneType Error"""
    try:
        return data.split(splitter)
    except:
        return ""

def status_position(column1, column2):
    # type(column) = float
    status = np.array([column1, column2])
    return status