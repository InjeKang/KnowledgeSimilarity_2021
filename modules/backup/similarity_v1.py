import pandas as pd
from os import path
import os
import openpyxl #https://stackoverflow.com/questions/54024504/reading-and-passing-excel-filename-with-pandas
import math


def read_xlsx(data):
    os.chdir("D:/Analysis/2021_Similarity/data/patent/KISTI")
    patent = openpyxl.load_workbook(data) # can no longer open xlsx file from pandas' read_excel >> used openpyxl
    if patent.sheetnames == ["Sheet0"]:
        sheet = patent["Sheet0"]
    elif patent.sheetnames == ["Sheet1"]:
        sheet = patent["Sheet1"]
    sheet.delete_rows(0, 2)
    firm_patent = pd.DataFrame(sheet.values)
    firm_patent = firm_patent[[9, 10, 32]] #select columns - application number, application year, ipc
    os.chdir("D:/Analysis/2021_Similarity/data")
    return firm_patent    

def split_column_into_lists(data, column, splitter):
    """Turn every rows of the column into list of lists"""
    data2 = data
    column2 = column.copy()
    column2list = column2.values.tolist() # series to list
    
    for i in range(len(column2list)):
        if column2list[i] is not None:
            column2list[i] = column2list[i].split(splitter)
            for j in range(len(column2list[i])):
                column2list[i][j] = column2list[i][j].strip()
        else:
            column2list[i] = ""

    # previous version below... did not work for certain cases (e.g. Acreo)...dont know why
    # for i in range(len(column2)):
    #     column2.iloc[i] = str(column2.iloc[i]).split(splitter)
    #     for j in range(len(column2.iloc[i])):
    #         column2.iloc[i][j] = column2.iloc[i][j].strip()
    column2list = pd.Series(column2list)
    return column2list

def column_to_list(data, column):
    """Make rows of a column into one list"""
    data2 = data
    firm_patent_list = []
    patent_list = split_column_into_lists(data2, column, ";")
    for i in range(len(column)):
        firm_patent_list.extend(patent_list.tolist()[i])
    # or 
    # firm_patent_list = [item for patent_list in t for item in firm_patent_list]
    return firm_patent_list

def create_placeholder_list(list_of_list):
    """Create a placeholder list (a) to prevent error of ind out of range and
    (b) to have a word (instead of an alphabet) as a string"""
    if isinstance(list_of_list, list):
        placeholder_list = []
        for i, _ in enumerate(list_of_list):
            placeholder_list_row = []
            for j, _ in enumerate(list_of_list[i]):
                placeholder_list_row.append("")
            placeholder_list.append(placeholder_list_row)
        return placeholder_list
    else:
        placeholder_list = []
        for i, _ in enumerate(list_of_list):
            placeholder_list_row = []
            for j, _ in enumerate(list_of_list.iloc[i]):
                placeholder_list_row.append("")
            placeholder_list.append(placeholder_list_row)
        return placeholder_list

def jaffe_similarity(focal_firm, partner_firm, focal, partner, unique): # ..._patent_list
    """measuring technological similarity"""
    focal_firm = focal_firm
    partner_firm = partner_firm
    focal_list = focal
    partner_list = partner
    unique_list = unique
    unique_focal = create_placeholder_list(unique_list)
    unique_partner = create_placeholder_list(unique_list)
    #remove whitespace to unify the format
    for x in range(len(unique_list)):
        unique_list[x] = unique_list[x].replace(" ", "")
    for y in range(len(focal_list)):
        focal_list[y] = focal_list[y].replace(" ", "")
    for z in range(len(partner_list)):
        partner_list[z] = partner_list[z].replace(" ", "")    
    # fraction of a focal and partner firm   
    for i in range(len(unique_list)):
        unique_focal[i] = 0
        unique_partner[i] = 0
        for j in range(len(focal_list)):
            if unique_list[i] == focal_list[j]:
                unique_focal[i] = unique_focal[i] + 1
        for k in range(len(partner_list)):
            if unique_list[i] == partner_list[k]:
                unique_partner[i] = unique_partner[i] + 1
    # measuring similarity
    if len(focal_list)*len(partner_list) != 0: # in case when a focal or partner firm does not have patents in the designated period
        unique_focal = [i/len(focal_list) for i in unique_focal]
        unique_partner = [i/len(partner_list) for i in unique_partner]
        focalXpartner = [a*b for a,b in zip(unique_focal, unique_partner)]
        unique_focal2 = [i**2 for i in unique_focal]
        unique_partner2 = [i**2 for i in unique_partner]
        Jaffesimilarity = sum(focalXpartner) / (math.sqrt(sum(unique_focal2)) * math.sqrt(sum(unique_partner2)))
    else:
        Jaffesimilarity = ""
    return Jaffesimilarity

def measure_similarity(data):
    """making a variable of technological similarity"""
    tech_similarity = []
    alliance_firm = []
    for i in range(len(data)):    
        # loading a focal firm        
        for focal_name in os.listdir("./patent/KISTI/"):
            if focal_name.lower().startswith(data["focal"][i].lower()): # read a focal firm's patents
                focal_firm=""
                focalFirm = focal_firm + focal_name            
                focalFirm_patent = read_xlsx(focalFirm)          
                year = int(data["year"][i])
                focal_patent = focalFirm_patent[(pd.to_numeric(focalFirm_patent[10]) >= year-5) &
                (pd.to_numeric(focalFirm_patent[10]) <= year-1)].reset_index(drop=True, inplace=False)  # select patents applied in between t-5 ~ t-1
                focal_patent = focal_patent.drop_duplicates(subset = [9]).reset_index(drop=True, inplace=False) 
                focal_patent_list = []
                focal_patent_list = column_to_list(focalFirm, focal_patent[32])
                # loading a partner firm
                for partner_name in os.listdir("./patent/KISTI/"):    
                    if partner_name.lower().startswith(data["partner"][i].lower()): # read a partner firm's patenets
                        partner_firm=""
                        partnerFirm = partner_firm + partner_name            
                        partnerFirm_patent = read_xlsx(partnerFirm)
                        partner_patent = partnerFirm_patent[(pd.to_numeric(partnerFirm_patent[10]) >= year-5) & 
                        (pd.to_numeric(partnerFirm_patent[10]) <= year-1)].reset_index(drop=True, inplace=False)  # select patents applied in between t-5 ~ t-1
                        partner_patent = partner_patent.drop_duplicates(subset = [9]).reset_index(drop=True, inplace=False) 
                        partner_patent_list = []
                        partner_patent_list = column_to_list(partnerFirm, partner_patent[32])
                        # making a unique patent list
                        unique_patent_list = []
                        unique_patent_list = focal_patent_list + partner_patent_list
                        unique_patent_list = list(set(unique_patent_list))
                        # measuring similarity between a focal and partner firm
                        similarity_result = jaffe_similarity(focalFirm, partnerFirm, focal_patent_list, partner_patent_list, unique_patent_list)
                        tech_similarity.append([similarity_result])
                        # tech_similarity = [x for x in tech_similarity if str(x) != "na"] # remove 'na' >> should not remove because of length
                        partners = focalFirm + "-" + partnerFirm
                        alliance_firm.append(partners)                        
        print("measuring similarity # {}".format(i+1))
    data["similarity"] = tech_similarity
    return data


# checking errors

# # for i in range(len(alliance_firm)):
# #     print(alliance_firm[i], i+1)

# # import pandas as pd

# # firm = pd.read_excel("D:/Analysis/2021_Similarity/data/02.alliance_v3(firms_with_patents_v2).xlsx", engine="openpyxl")
# # focal = firm["focal"]+".xlsx-"
# # partner = firm["partner"]+".xlsx"
# # focal_partner = pd.concat([focal, partner], axis=1)
# # focal_partner["focal_partner"] = focal_partner["focal"] + focal_partner["partner"]

# # for i in range(718, 1119):
# #     if alliance_firm[i].lower() == focal_partner["focal_partner"][i-4].lower():
# #         pass
# #     else:
# #         print("error! # {}".format(i))
# #         print(alliance_firm[i])
# #         break

# # for x in range(713, 718):
# #     print(alliance_firm[x])

# if __name__ == "__main__":
#     os.chdir(
#         path.join(os.getcwd(), "data")
#     ) 
#     raw_data = pd.read_excel("02.alliance_v3(firms_with_patents_v2).xlsx", engine="openpyxl", sheet_name="alliance")
#     result = measure_similarity(raw_data)