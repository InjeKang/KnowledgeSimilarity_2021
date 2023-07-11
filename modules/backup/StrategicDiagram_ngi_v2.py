import pandas as pd
import os
import openpyxl
import glob
from tqdm import trange
import math
from modules import lookup
from multiprocessing import Pool, cpu_count
import numpy as np

def read_xlsx_glob(data):    
    patent = openpyxl.load_workbook(data) # can no longer open xlsx file from pandas' read_excel >> used openpyxl
    if patent.sheetnames == ["Sheet0"]:
        sheet = patent["Sheet0"]
    elif patent.sheetnames == ["Sheet1"]:
        sheet = patent["Sheet1"]
    sheet.delete_rows(0, 2)
    firm_patent = pd.DataFrame(sheet.values)
    firm_patent = firm_patent[[9, 10, 18, 32]] #select columns - application number, application year, assignee, ipc
    # os.chdir("D:/Analysis/2021_Similarity/data")
    return firm_patent

def patent_stock(data): # data: alliance data
    """ accumulating patent data applied during t-5 and t-1"""
    # read all patent data into one dataframe
    # excel_files = glob.glob("D:/Analysis/2021_Similarity/data/patent/KISTI*.xlsx")
    data2 = data.copy()
    # excel_files = glob.glob("D:/Analysis/2021_Similarity" + "\*.xlsx") # for pilot test...524 data when using five companies
    excel_files = glob.glob("D:/Analysis/2021_Similarity/data/patent/KISTI" + "\*.xlsx")
    df = pd.concat((read_xlsx_glob(files) for files in excel_files))
    # select patent to make columns for StrategicDiagram
    patent_stock_list_year = []
    for i in trange(len(data2)):
        # select patent applied during the designated period
        year = int(data2["year"][i])
        patent_data = df[(pd.to_numeric(df[10]) >= year-5) & 
        (pd.to_numeric(df[10]) <= year-1)].reset_index(drop=True, inplace=False)         
        patent_data = patent_data.drop_duplicates(subset = [9]).reset_index(drop=True, inplace=False) 
        # make a list of assigned years
        patent_data_year = [patent_data[10].values.tolist()]
        patent_stock_list_year.extend(patent_data_year)
        print("accumulating patent stock # {}".format(i+1))
    # add the columns >> assigness/ipc of patents applied during the designated period
    data2["patent_stock_year"] = patent_stock_list_year    
    return data2

def ngi_firm(firm, firm_year, type_): # type_ = focal OR partner
    """
    a function to measure NGI of firms
    """
    df = firm_year
    # linking each firm with patent assigned year
    firm_list = []
    year_list = []
    for i in range(len(df)):
        if len(df["firm"][i]) == 1:
            firm_list.append("".join(df["firm"][i]))
            year_list.append(int(df["year"][i]))
        else:
            for j in range(len(df["firm"][i])):
                firm_list.append("".join(df["firm"][i][j]))
                year_list.append(int(df["year"][i]))
    # unify firms' names
    firm_list2 = lookup.unify_firm_name(firm_list, 1)
    firm_list2 = lookup.unify_firm_name2(firm_list2)
    df2 = pd.DataFrame((zip(firm_list2, year_list)), columns = ["firm", "year"])

    # measuring GI of every firms
    df2_mean = pd.Series(df2.groupby("firm")["year"].mean(), name = "mean")
    df2_min = pd.Series(df2.groupby("firm")["year"].min(), name = "min")
    df2_max = pd.Series(df2.groupby("firm")["year"].max(), name = "max")
    df3 = pd.concat([df2_mean, df2_min, df2_max], axis = 1)
    df3 = df3.drop(labels = "", axis = 0)
    df3["firm"]  = df3.index
    df3["gi"] = df3.apply(lambda x: ((x["mean"] - x["min"]) / (x["max"] - x["min"]))
    if x["max"] - x["min"] != 0 else 0, axis = 1)

    # measuring NGI of every firms
    gi_mean = df3["gi"].mean()
    gi_sd = df3["gi"].std()
    df3["ngi"] = df3["gi"].map(lambda x: (x-gi_mean)/gi_sd)

    # selecting GI of a focal/partner firm
    if type_ == "focal":
        focal = firm.lower()
        ngi_result = df3.loc[(df3["firm"].str.contains(focal))].reset_index(drop=True, inplace=False)
    else:
        partner = firm.lower()
        ngi_result = df3.loc[(df3["firm"].str.contains(partner))].reset_index(drop=True, inplace=False)        

    # check if the result is a unique value and measure NPI
    if len(ngi_result) == 1:
        # measuring NGI of every firms
        ngi_result2 = ngi_result["ngi"][0]
    else:
        print("array format")
        print(ngi_result)
        if ngi_result.empty:
            ngi_result2 = ""

        # elif gi_result.empty:
        #     ngi_result = ""
        # else:
        #     print(gi_result)
        #     sum_ = gi_result["ngi"].sum()
        #     ngi_result = ((sum_ - gi_mean) / gi_sd)
        #     print(ngi_result)
            
    return ngi_result2

  
def ngiFirm_to_alliance(data):
    """add NGI of firm to alliance data"""
    data2 = data.copy()    
    firm_list = data2["patent_stock_firm"].copy()
    year_list = data2["patent_stock_year"].copy()
    ngi_focal = []
    ngi_partner = []
    for i in trange(len(data2)):
        # preliminary work to select patents applied by a focal/partner firm
        df_firm_year = pd.DataFrame((zip(firm_list[i], year_list[i])), columns = ["firm", "year"])        
        # ngi of a focal firm
        focal = ngi_firm(data2["focal"][i], df_firm_year, "focal")
        ngi_focal.append(focal)
        # ngi of a partner firm
        partner = ngi_firm(data2["partner"][i], df_firm_year, "partner")
        ngi_partner.append(partner)
        print("measuring NGI # {}".format(i+1))

    # make columns
    data2["ngi_focal"] = ngi_focal
    data2["ngi_partner"] = ngi_partner

    return data2

def split_(data, splitter):
    """to split string with the consideration of NoneType Error"""
    try:
        return data.split(splitter)
    except:
        return ""


# def ipc_with_year(df):
#     ipc_list = []    
#     year_list = []    
#     for i in range(len(df)):
#         if len(df["ipc"][i]) == 0: #in case when there was NoneType above
#             pass
#         elif len(df["ipc"][i]) == 1:
#             ipc_list_ = list(x.replace(" ", "") for x in df["ipc"][i]) # to remove whitespace
#             ipc_list.append("".join(ipc_list_))
#             year_list.append(int(df["year"][i]))        
#         else:
#             for j in range(len(df["ipc"][i])):
#                 ipc_list_ = list((x.replace(" ", "") for x in df["ipc"][i])) # to remove whitespace
#                 ipc_list.append("".join(ipc_list_[j]))
#                 year_list.append(int(df["year"][i]))


def ngi_ipc(ipc_firm, ipc_year):
    """
    a function to measure NGI of IPCs
    """
    df = ipc_year.copy()
    ipc_list = []    
    year_list = [] 
    # linking each ipc with patent assigned year
    df["ipc"] = df["ipc"].apply(lambda x: split_(x, ";")) 
    for i in range(len(df)):
        if len(df["ipc"][i]) == 0: #in case when there was NoneType above
            pass
        elif len(df["ipc"][i]) == 1:
            ipc_list_ = list(x.replace(" ", "") for x in df["ipc"][i]) # to remove whitespace
            ipc_list.append("".join(ipc_list_))
            year_list.append(int(df["year"][i]))        
        else:
            for j in range(len(df["ipc"][i])):
                ipc_list_ = list((x.replace(" ", "") for x in df["ipc"][i])) # to remove whitespace
                ipc_list.append("".join(ipc_list_[j]))
                year_list.append(int(df["year"][i]))
    df2 = pd.DataFrame((zip(ipc_list, year_list)), columns = ["ipc", "year"])

    # measuring GI of every IPC
    df2_mean = pd.Series(df2.groupby("ipc")["year"].mean(), name = "mean")
    df2_min = pd.Series(df2.groupby("ipc")["year"].min(), name = "min")
    df2_max = pd.Series(df2.groupby("ipc")["year"].max(), name = "max")
    df3 = pd.concat([df2_mean, df2_min, df2_max], axis = 1)
    try:
        df3 = df3.drop(labels = "", axis = 0) 
    except: # "[''] not found in axis"
        pass
    df3["ipc"]  = df3.index
    df3["gi"] = df3.apply(lambda x: ((x["mean"] - x["min"]) / (x["max"] - x["min"]))
    if x["max"] - x["min"] != 0 else 0, axis = 1)

    # measuring NGI of every firms
    gi_mean = df3["gi"].mean()
    gi_sd = df3["gi"].std()
    df3["ngi"] = df3["gi"].map(lambda x: (x-gi_mean)/gi_sd)

    # selecing NGIs of non-overlapping IPCs of focal/partner firm
    ngi_result= df3[df3.apply(lambda x: any([y in x["ipc"] for y in ipc_firm]), axis = 1)].reset_index(drop=True, inplace=False)

    # average of the selected NPIs
    ngi_result2 = ngi_result["ngi"].mean()
            
    return ngi_result2

def ngiIPC_to_alliance(data):
    """add NGI of IPC to alliance data"""
    data2 = data.copy()
    ipc_list = data2["patent_stock_ipc"].copy()    
    year_list = data2["patent_stock_year"].copy()
    ngi_focal = []
    ngi_partner = []
    for i in trange(len(data2)):
    # for i in range(3, 4):
        # preliminary work to select patents applied by a focal/partner firm
        df_ipc_year = pd.DataFrame((zip(ipc_list[i], year_list[i])), columns = ["ipc", "year"])        
        # ngi of a focal firm
        focal_ipc_list = list((x.replace(" ", "") for x in data2["focal_ipc"][i])) # to remove white space
        focal = ngi_ipc(focal_ipc_list, df_ipc_year)
        ngi_focal.append(focal)        
        # ngi of a partner firm        
        partner_ipc_list = list((x.replace(" ", "") for x in data2["partner_ipc"][i]))
        partner = ngi_ipc(partner_ipc_list, df_ipc_year)
        ngi_partner.append(partner)
        print("measuring NGI # {}".format(i+1))

    # make columns
    data2["ngi_ipc_focal"] = ngi_focal
    data2["ngi_ipc_partner"] = ngi_partner

    return data2

# df = df_ipc_year.copy()
# for i in range(len(df)):
#     # if df["ipc"][i] is None:
#     #     print(i)
#     #     print(df["ipc"][i])
#     #     break
#     try:
#         df["ipc"][i] = df["ipc"][i].split(";")
#     except:
#         print(i)
#         print(df["ipc"][i])
#         break #118193

