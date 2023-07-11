import pandas as pd
import os
import openpyxl
import glob
from tqdm import trange
import math

def read_xlsx_glob(data):    
    patent = openpyxl.load_workbook(data) # can no longer open xlsx file from pandas' read_excel >> used openpyxl
    if patent.sheetnames == ["Sheet0"]:
        sheet = patent["Sheet0"]
    elif patent.sheetnames == ["Sheet1"]:
        sheet = patent["Sheet1"]
    sheet.delete_rows(0, 2)
    firm_patent = pd.DataFrame(sheet.values)
    firm_patent = firm_patent[[9, 10, 18, 32]] #select columns - application number, application year, assignee, ipc
    # os.chdir("D:/Analysis/2021_Similarity/data")
    return firm_patent

def patent_stock(data): # data: alliance data
    """ accumulating patent data applied during t-5 and t-1"""
    # read all patent data into one dataframe
    # excel_files = glob.glob("D:/Analysis/2021_Similarity/data/patent/KISTI*.xlsx")
    data2 = data.copy()
    # excel_files = glob.glob("D:/Analysis/2021_Similarity" + "\*.xlsx") # for pilot test...524 data when using five companies
    excel_files = glob.glob("D:/Analysis/2021_Similarity/data/patent/KISTI" + "\*.xlsx")
    df = pd.concat((read_xlsx_glob(files) for files in excel_files))
    # select patent to make columns for StrategicDiagram
    patent_stock_list_year = []
    for i in trange(len(data2)):
        # select patent applied during the designated period
        year = int(data2["year"][i])
        patent_data = df[(pd.to_numeric(df[10]) >= year-5) & 
        (pd.to_numeric(df[10]) <= year-1)].reset_index(drop=True, inplace=False)         
        patent_data = patent_data.drop_duplicates(subset = [9]).reset_index(drop=True, inplace=False) 
        # make a list of assigned years
        patent_data_year = [patent_data[10].values.tolist()]
        patent_stock_list_year.extend(patent_data_year)
        print("accumulating patent stock # {}".format(i+1))
    # add the columns >> assigness/ipc of patents applied during the designated period
    data2["patent_stock_year"] = patent_stock_list_year    
    return data2

def ngi_firm(firm, firm_year, type_): # type_ = focal OR partner
    """
    a function to measure NGI of firms
    """
    df = firm_year
    # linking each firm with patent assigned year
    firm_list = []
    year_list = []
    for i in range(len(df)):
        if len(df["firm"][i]) == 1:
            firm_list.append("".join(df["firm"][i]))
            year_list.append(int(df["year"][i]))
        else:
            for j in range(len(df["firm"][i])):
                firm_list.append("".join(df["firm"][i][j]))
                year_list.append(int(df["year"][i]))
    df2 = pd.DataFrame((zip(firm_list, year_list)), columns = ["firm", "year"])    
    # measuring GI of every firms
    df2_mean = pd.Series(df2.groupby("firm")["year"].mean(), name = "mean")
    df2_min = pd.Series(df2.groupby("firm")["year"].min(), name = "min")
    df2_max = pd.Series(df2.groupby("firm")["year"].max(), name = "max")
    df3 = pd.concat([df2_mean, df2_min, df2_max], axis = 1)
    df3 = df3.drop(labels = "", axis = 0)
    df3["firm"]  = df3.index
    # selecting GI of a focal/partner firm
    if type_ == "focal":
        focal = firm.lower()
        gi_result = df3.loc[(df3["firm"].str.contains(focal))].reset_index(drop=True, inplace=False)
    else:
        partner = firm.lower()
        gi_result = df3.loc[(df3["firm"].str.contains(partner))].reset_index(drop=True, inplace=False)
        
    
    
    df3["gi"] = df3.apply(lambda x: ((x["mean"] - x["min"]) / (x["max"] - x["min"]))
    if x["max"] - x["min"] != 0 else 0, axis = 1)
    # selecting GI of a focal/partner firm
    if type_ == "focal":
        focal = firm.lower()
        gi_result = df3.loc[(df3["firm"].str.contains(focal))].reset_index(drop=True, inplace=False)
    else:
        partner = firm.lower()
        gi_result = df3.loc[(df3["firm"].str.contains(partner))].reset_index(drop=True, inplace=False)
    # check if the result is a unique value and measure NPI
    if len(gi_result) == 1:
        # measuring NGI of every firms
        gi_mean = df3["gi"].mean()
        gi_sd = df3["gi"].std()
        df3["ngi"] = df3["gi"].map(lambda x: (x-gi_mean)/gi_sd)
        ngi_result = gi_result["npi"][0]
        print("pass")
    elif gi_result.empty:
        ngi_result = ""
    else:
        
    # else:
    #     if ngi_result.empty:
    #         ngi_result2 = ""
    #     else:
    #         sum_ = math.log(ngi_result["npi"].sum())
    #         ngi_result2 = sum_ - gi_mean*(len(ngi_result) -1 )
    #         print(sum_)
    
    return ngi_result2



            
            

        

def ngiFirm_to_alliance(data):
    """add NGI of firm to alliance data"""
    data2 = data.copy()    
    firm_list = data2["patent_stock_firm"].copy()
    year_list = data2["patent_stock_year"].copy()
    for i in trange(len(data2)):
        # preliminary work to select patents applied by a focal/partner firm
        df_firm_year = pd.DataFrame((zip(firm_list[i], year_list[i])), columns = ["firm", "year"])        
        # ngi of a focal firm
        focal = ngi_firm(data2["focal"][i], df_firm_year, "focal")


    
    ngi_focal = []
    ngi_partner = []
    # for i in trange(len(data2)):
        