from tracemalloc import stop
from os import path
from tqdm import trange
from tqdm import tqdm
from modules import lookup
import pandas as pd
import os
import openpyxl #https://stackoverflow.com/questions/54024504/reading-and-passing-excel-filename-with-pandas
import math
import statistics
import copy
import glob
import re
import collections

def read_xlsx_glob(data):    
    patent = openpyxl.load_workbook(data) # can no longer open xlsx file from pandas' read_excel >> used openpyxl
    if patent.sheetnames == ["Sheet0"]:
        sheet = patent["Sheet0"]
    elif patent.sheetnames == ["Sheet1"]:
        sheet = patent["Sheet1"]
    sheet.delete_rows(0, 2)
    firm_patent = pd.DataFrame(sheet.values)
    firm_patent = firm_patent[[9, 10, 18, 32]] #select columns - application number, application year, assignee, ipc
    # os.chdir("D:/Analysis/2021_Similarity/data")
    return firm_patent    

def split_column_into_lists(column, splitter):
    """Turn each columns into list of lists"""
    column = column.copy()
    for i in range(len(column)):
        column.iloc[i] = str(column.iloc[i]).split(splitter)
        for j in range(len(column.iloc[i])):
            column.iloc[i][j] = column.iloc[i][j].strip()
    return column

def column_to_list(column):
    """Make rows of a column into a list"""
    firm_patent_list = []
    patent_list = split_column_into_lists(column, ";")
    for i in range(len(column)):
        firm_patent_list.extend(patent_list.tolist()[i])
    return firm_patent_list

def create_placeholder_list(list_of_list):
    """Create a placeholder list (a) to prevent error of index out of range and
    (b) to have a word (instead of an alphabet) as a string"""
    if isinstance(list_of_list, list):
        placeholder_list = []
        for i, _ in enumerate(list_of_list):
            placeholder_list_row = []
            for j, _ in enumerate(list_of_list[i]):
                placeholder_list_row.append("")
            placeholder_list.append(placeholder_list_row)
        return placeholder_list
    else:
        placeholder_list = []
        for i, _ in enumerate(list_of_list):
            placeholder_list_row = []
            for j, _ in enumerate(list_of_list.iloc[i]):
                placeholder_list_row.append("")
            placeholder_list.append(placeholder_list_row)
        return placeholder_list

def affil_country(data):
    affiliation = data["patent_stock_firm"].copy()
    affiliation_splitted = split_column_into_lists(affiliation, ";")
    placeholder_list = create_placeholder_list(affiliation_splitted)
    # create an empty placeholder list
    country_list = copy.deepcopy(placeholder_list)
    # Selecting countries from the column
    for i in range(len(data)):
        for j in range(len(affiliation_splitted[i])):      
            try:
                country_list[i][j] = affiliation_splitted[i][j].split("`")[0]
            except:
                country_list[i][j] = ""
    return country_list
 
def affil_firm(data):
    data2 = data.copy()
    # data2 = patent_data.copy() # for test
    affiliation = data2[18].copy()
    affiliation = affiliation.str.replace(" ", "")
    affiliation_splitted = split_column_into_lists(affiliation, ";")
    placeholder_list = create_placeholder_list(affiliation_splitted)
    # create an empty placeholder list
    firm_list = copy.deepcopy(placeholder_list)
    # Selecting countries from the column
    stopwords = ["corporate", "corporation", "us", "fr", "company", "corporate", "corp"]
    for i in range(len(data2)):
        for j in range(len(affiliation_splitted[i])):    
            try:
                # [-2:-1] because there are some data that ends with ' '
                # while others ends with 'corp_name'
                firm_list[i][j] = affiliation_splitted[i][j].split("`")[-5:]
                # firm_list[i][j] = affiliation_splitted[i][j].split("``")
                # remove punctuations or unnecessary words
                firm_list[i][j] = " ".join(firm_list[i][j]) # list to string
                firm_list[i][j] = firm_list[i][j].lower()
                firm_list[i][j] = firm_list[i][j].replace("`", " ")
                firm_list[i][j] = re.sub(r'[^\w\s]','',firm_list[i][j])
                firm_list[i][j] = [word for word in firm_list[i][j].split() if word not in stopwords]
                firm_list[i][j] = list(set(firm_list[i][j]))
                firm_list[i][j] = lookup.unify_firm_name(firm_list[i][j], 0)
                firm_list[i][j] = ", ".join(firm_list[i][j]) # list to string                                
                # for test, 0, 3505, 100, 3858, 408
            except:
                firm_list[i][j] = ""    
        # firm_list[i] = [w.replace(" ", "") for w in firm_list[i]]           
        # firm_list[i] = [x for x in firm_list[i] if x]    
    # data2["firm_cleansed"] = firm_list
    return firm_list

def patent_stock(data): # data: alliance data
    """ accumulating patent data applied during t-5 and t-1"""
    # read all patent data into one dataframe
    # excel_files = glob.glob("D:/Analysis/2021_Similarity/data/patent/KISTI*.xlsx")
    data2 = data.copy()
    # excel_files = glob.glob("D:/Analysis/2021_Similarity" + "\*.xlsx") # for pilot test...524 data when using five companies
    excel_files = glob.glob("D:/Analysis/2021_Similarity/data/patent/KISTI" + "\*.xlsx")
    df = pd.concat((read_xlsx_glob(files) for files in excel_files))
    # select patent to make columns for StrategicDiagram
    patent_stock_list_firm = []
    patent_stock_list_ipc = []
    for i in range(len(data2)):
        # select patent applied during the designated period
        year = int(data2["year"][i])
        patent_data = df[(pd.to_numeric(df[10]) >= year-5) & 
        (pd.to_numeric(df[10]) <= year-1)].reset_index(drop=True, inplace=False) 
        patent_data = patent_data.drop_duplicates(subset = [9]).reset_index(drop=True, inplace=False) 
        # make a list of assignees(18)
        ## cleansing firm data
        # patent_data_cleansed = affil_firm(patent_data)
        # patent_data_firm = [patent_data_cleansed["firm_cleansed"].tolist()]
        patent_data_firm = [affil_firm(patent_data)]
        patent_stock_list_firm.extend(patent_data_firm)        
        # make a list of ipc(32)
        patent_data_ipc = [patent_data[32].values.tolist()]
        patent_stock_list_ipc.extend(patent_data_ipc)
        print("accumulating patent stock # {}".format(i+1))
    # add the columns >> assigness/ipc of patents applied during the designated period
    data2["patent_stock_firm"] = patent_stock_list_firm
    data2["patent_stock_ipc"] = patent_stock_list_ipc
    return data2

def npi_firm(firm, list_full, list_unique, type): # type = focal OR partner
    """
    a function to measure NPI
    """
    #1 measuirng NPI of a focal firm
    #2 making a dataframe to match total production for each firm
    # tp = [0] * len(list_unique)
    # df = pd.DataFrame(zip(list_unique, tp), columns = ["firm", "tp"])

    #2.1 measuring total production of each firm
    vals = collections.Counter(list_full).values()
    keys = collections.Counter(list_full).keys()
    df = pd.DataFrame({"firm": keys, "tp": vals})

    # for i in range(len(df)):
    #     if df["firm"][i] in list_full: ## this part error
    #         df["tp"][i] += 1
    #2.1 measuring total production of each firm (alternative 1)
    """
    when checking with "acreo", the first version's df["tp"] is 1
    while, the other version is 57182
    ...seems the former fails to count 
    """
    # for i in range(len(df)):
    #     for j in range(len(list_full)):
    #         if df["firm"][i] == list_full[j]:
    #             df["tp"][i] += 1    
    # #2.1 measuring total production of each firm (alternative 2)
    # for i in range(len(df)):
    #     df["tp"][i] = sum(x in df["firm"][i] for x in list_full)

    #3 measuring npi of firms
    df["log_tp"] = df["tp"].map(lambda x: math.log(x)) # map can be used in Series...progress_map/process_map does not work
    tp_mean = df["log_tp"].mean() # https://towardsdatascience.com/dplyr-style-data-manipulation-with-pipes-in-python-380dcb137000
    tp_sd = df["log_tp"].std()
    df["npi"] = df["log_tp"].map(lambda x: (x-tp_mean)/tp_sd)
    #3.1 selecting npi of a focal/partner firm    
    if type == "focal":
        focal = firm.lower()
        npi_result = df.loc[(df["firm"].str.contains(focal))].reset_index(drop=True, inplace=False)        
    else:
        partner = firm.lower()
        npi_result = df.loc[(df["firm"].str.contains(partner))].reset_index(drop=True, inplace=False)
    #3.2 check if there is a unique value
    if len(npi_result) == 1:    
        npi_result2 = npi_result.iloc[0]["npi"]
        print("pass") # no meaning of using "i" >> remove when re-run
    else: # checked if they are the same company
        print("array format") # no meaning of using "i" >> remove when re-run
        print(npi_result)
        # add the same firm with different names
        if npi_result.empty:
            npi_result2 = ""
        else:
            sum_ = math.log(npi_result["tp"].sum())
            npi_result2 = ((sum_ - tp_mean)/tp_sd)
            print(sum_)
            print(npi_result2)
            

    # if type(npi_result.iloc[0]["npi"]) == str:    # why did I make this code...
    #     npi_result2 = npi_result.iloc[0]["npi"]
    # else:
    #     print("array format in # {}".format(i))
    #     print(firm.lower())

    return npi_result2

# for i in range(len(firm_list[0])):
#     if " ".join(firm_list[0][i]).find("kog") >= 0:
#         print(i)



def measure_npi_firm(data):
    """measuring Normalized Performance Index"""
    data2 = data.copy()
    # data2 = patent_stock(data2)
    firm_list = data2["patent_stock_firm"].copy()    
    npi_focal = []
    npi_partner = []    
    for i in trange(len(data2)):            
        firm_list_full = []
        #1 making a unique firm list
        for j in range(len(firm_list[i])):
            if len(firm_list[i][j]) == 1:
                list_to_string = "".join(firm_list[i][j])
                firm_list_full.append(list_to_string)
            else:
                for k in range(len(firm_list[i][j])):
                    list_to_string = "".join(firm_list[i][j][k])
                    firm_list_full.append(list_to_string)
        firm_list_full = lookup.unify_firm_name(firm_list_full, 1)
        firm_list_unique = list(set(firm_list_full))
        firm_list_unique = [x for x in firm_list_unique if x] # remove empty string
    #2 npi of a focal firm
        focal = npi_firm(data2["focal"][i], firm_list_full, firm_list_unique, "focal")
        npi_focal.append(focal)
    #3 npi of a partner firm
        partner = npi_firm(data2["partner"][i], firm_list_full, firm_list_unique, "partner")
        npi_partner.append(partner) #add print(i) after this line
        print("measuring NPI # {}".format(i+1))
    #4 make columns        
    data2["npi_focal"] = npi_focal
    data2["npi_partner"] = npi_partner
    return data2

if __name__ == "__main__":
    os.chdir(
        path.join(os.getcwd(), "data")
    ) 
    raw_data = pd.read_excel("03.Result_v1(20220405).xlsx", engine="openpyxl", sheet_name="Sheet1").drop("Unnamed: 0", axis=1)
    result = measure_npi_firm(
        patent_stock(raw_data)
        )

        